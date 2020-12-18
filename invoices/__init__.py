import os
import sys

import openpyxl
import datetime
import sqlite3
from shutil import copyfile
import glob
from dateutil import parser
import smtplib


from email.utils import COMMASPACE, formatdate
from email.mime.text import MIMEText
from InvoiceGenerator.api import Invoice, Item, Client, Provider, Creator
from InvoiceGenerator.pdf import SimpleInvoice
from common import Common
from common.gdrive import GDrive

c = Common()
gdrive=GDrive()
# choose english as language
os.environ["INVOICE_LANG"] = "en"

creator = Creator('Life in Harmony @ VCC Admin')


dict = {'Dave': 'D K', 'Stefanie': 'S F', 'Sabrina': 'S G',
        'Mariella': 'M P', 'Johanna': 'J G', 'Sandra': 'S M',
        'Meri': 'M C', 'Meri': 'M P', 'Teresa': 'T C',
        'Peter': 'P M', 'Rosemary': 'R C',
        'Alex': 'A E',
        'Mary': 'Mary D', 'Kelly-Lee': 'K', 'Lisa': 'Lisa B',
        'Joanne': 'Joanne D', 'Helena':'H'}

myDictionary = {}
superDictionary = {}


class MyInvoice(SimpleInvoice):
    def __init__(self, invoice):
        super(MyInvoice, self).__init__(invoice)
    def _drawTitle(self):
        # Up line
        self.pdf.drawString(self.LEFT*2.834, self.TOP*2.834, self.invoice.title)
        self.pdf.drawString(
            (self.LEFT + 90) * 2.834,
            self.TOP*2.834,
            _(u'Associate-Client Summary: %s') % self.invoice.number,
        )

def analyze(collector, payee, db, sDate, eDate, invoicePath, gfile):
    try:
        today = datetime.datetime.now().today()
        database = sqlite3.connect(db)
        database.row_factory = sqlite3.Row
        database.text_factory = str

        cursor = database.cursor()

        cursor.execute('''
                    CREATE TABLE IF NOT EXISTS associate_invoices(id INTEGER PRIMARY KEY, name TEXT, number TEXT unique,
                                       invoice_date DATETIME, tax REAL, subtotal REAL, total REAL, status TEXT, last_update DATETIME, invoice_file BLOB)
                ''')
        cursor.execute('''
                    CREATE TABLE IF NOT EXISTS supervisor_invoices(id INTEGER PRIMARY KEY, super_name TEXT, invoice_date DATETIME,
                                        number TEXT unique, tax REAL, subtotal REAL, total REAL, status TEXT, last_update DATETIME
                )''')
        database.commit()
        eDateInvoices = parser.parse(eDate)
        print 'Considering sales between ' + sDate + ' and ' + eDate

        cursor.execute('''SELECT * FROM invoices where name=? and invoice_date < ? and (service_type like "%Supervision%" or service_type like "%supervision%")''', (collector, eDateInvoices))
        allinvoices = cursor.fetchall()
        cursor2 = database.cursor()
        cursor5 = database.cursor()
        # writer.writeheader()
        # myworkbook = openpyxl.load_workbook(filename=processedFile)
        invoiceSW = None
        invoicePsy = None
        isPercentage = False

        if collector == 'Alex' or collector == 'Mary' or collector == 'Lisa' or collector == 'Helena':
            isPercentage = True

        # for row in inSheet.rows:
        invoiceid = None
        nowStr = str(today.strftime('%m/%d/%Y-%H%M%S'))

        for row in allinvoices:
            SWorPsy = None
            isPartialInvoice = False
            if row['is_partial_invoice'] == 1:
                isPartialInvoice = True

            if row['status'] == 'reversed':
                continue

            isNewSupervisorInvoice=False

            if row['service_type'] is not None and (str(row['service_type']).lower().find('psycholo') != -1):
                SWorPsy = 'psy'
            elif row['service_type'] is not None and (str(row['service_type']).lower().find('sw supervision') != -1 or str(row['service_type']).lower().find('rsw') != -1):
                SWorPsy = 'sw'
            else:
                continue

            collectorKey = collector + "-" +SWorPsy

            if row['supervisor_invoice_fk'] is None and row['status'] != 'cancelled':
                isNewSupervisorInvoice = True
                superInvoice=None
                if superDictionary.get(collectorKey) is not None:
                    cursor2.execute('''SELECT last_update from supervisor_invoices where id=?''',
                                    (superDictionary[collectorKey],))
                    superInvoice=cursor2.fetchone()
                if superInvoice is None:
                        cursor2.execute('''INSERT INTO supervisor_invoices(super_name, number, invoice_date, tax, subtotal, total, status, last_update, supervisor_type)
                                      VALUES(?,?,?,?,?,?,?,?,?)''',
                                        (collector, collector + "-" + SWorPsy + "-" + nowStr, eDateInvoices, 0, 0, 0, 'sent', today, SWorPsy.upper()))
                        superDictionary[collectorKey] = cursor2.lastrowid
                        invoiceid = superDictionary[collectorKey]

                cursor2.execute('''UPDATE invoices set supervisor_invoice_fk=? 
                        where number=?''', (superDictionary[collectorKey], row['number']))

                database.commit()
            else:
                cursor2.execute('''SELECT last_update from supervisor_invoices where id=?''',
                                (row['supervisor_invoice_fk'],))
                superInvoice = cursor2.fetchone()
                if superInvoice is not None:
                    a = parser.parse(superInvoice['last_update'])
                    b = parser.parse(row['last_update'])
                    if superInvoice is not None and a <= b and row['status'] == 'paid':
                        isNewSupervisorInvoice = True
                        print 'CHANGE CHANGE in status of ' + row['number']
                        if superDictionary.get(collectorKey) is None:
                            cursor2.execute('''INSERT INTO supervisor_invoices(super_name, number, invoice_date, tax, subtotal, total, status, last_update,  supervisor_type)
                                      VALUES(?,?,?,?,?,?,?,?,?)''',
                                            (collector, collector + "-" + SWorPsy + "-" + nowStr, eDateInvoices, 0, 0, 0, 'sent', today, SWorPsy.upper()))
                            superDictionary[collectorKey] = cursor2.lastrowid
                        invoiceid = superDictionary[collectorKey]
                        cursor2.execute('''UPDATE invoices set supervisor_invoice_fk=? where id=?''',
                                        (invoiceid, row['id']))
                        database.commit()

            if isNewSupervisorInvoice == False:
                continue

            client = Client(dict.get(collector))
            provider = Provider("Supervisor", bank_account=' ', bank_code=' ')

            cursor5.execute('''SELECT * FROM invoices where id=?''', (row['id'],))
            updatedRow = cursor5.fetchone()
            if SWorPsy == 'sw':
                invoiceSW=createInvoicePDF(client, invoiceSW, invoicePath, invoiceid, False, SWorPsy, False, provider, updatedRow, sDate, isPartialInvoice)
            else:
                invoicePsy=createInvoicePDF(client, invoicePsy, invoicePath, invoiceid, False, SWorPsy, False, provider, updatedRow, sDate, isPartialInvoice)

        if invoiceSW is not None and superDictionary[collector+"-sw"] is not None:
            cursor2.execute('''UPDATE supervisor_invoices set total=?, subtotal=?, tax=? 
                    where id=?''', (
            float(invoiceSW.price_tax), float(invoiceSW.price), float(invoiceSW.price_tax - invoiceSW.price), superDictionary[collector+"-sw"]))
            database.commit()
            print "total price sw is " + str(invoiceSW.price)
            print "total tax sw is " + str(invoiceSW.price_tax)

        if invoicePsy is not None and superDictionary[collector+"-psy"] is not None:
            cursor2.execute('''UPDATE supervisor_invoices set total=?, subtotal=?, tax=? 
                    where id=?''', (
            float(invoicePsy.price_tax), float(invoicePsy.price), float(invoicePsy.price_tax - invoicePsy.price), superDictionary[collector+"-psy"]))
            database.commit()
            print "total price psy is " + str(invoicePsy.price)
            print "total tax psy is " + str(invoicePsy.price_tax)

        #print "total price is "+str(invoiceSW.price)
        #print "total tax is "+str(invoiceSW.price_tax)

        if invoiceSW is not None:
            pdf_name = collector + "-SocialWorker-" + str(
                today.strftime("%Y%m%d-%H%M%S")) + ".pdf";
            pdf = MyInvoice(invoiceSW)
            pdfPath = invoicePath + "/" + pdf_name
            pdf.gen(pdfPath,
                    generate_qr_code=True)
            file = open(pdfPath, 'r')
            file_content = file.read()
            cursor2.execute('''UPDATE supervisor_invoices set invoice_file=? 
                    where id=?''', (file_content, superDictionary[collector+"-sw"]))
            database.commit()
            gdrive.create_file(gfile, pdf_name, pdfPath)
            file.close()


        if invoicePsy is not None:
            pdf = MyInvoice(invoicePsy)
            pdf_name = collector + "-Psychologist-" + str(
                today.strftime("%Y%m%d-%H%M%S")) + ".pdf"
            pdfPath = invoicePath + "/" + pdf_name
            pdf.gen(pdfPath,
                    generate_qr_code=True)

            file = open(pdfPath, 'r')
            file_content = file.read()
            cursor2.execute('''UPDATE supervisor_invoices set invoice_file=? 
                    where id=?''', (file_content, superDictionary[collector+"-psy"]))
            database.commit()
            gdrive.create_file(gfile, pdf_name, pdfPath)
            file.close()



        cursor3=database.cursor()
        cursor4=database.cursor()
        cursor3.execute('''SELECT * FROM invoices where name=? and invoice_date < ?''', (collector, eDateInvoices))
        allinvoices = cursor3.fetchall()

        associate_invoice = None
        invoiceid = None
        invoice = None
        for row in allinvoices:
            isNewOrChangedClientInvoice = False

            isPartialInvoice = False
            if row['is_partial_invoice'] == 1:
                isPartialInvoice = True

            if row['status'] == 'reversed':
                continue

            isRental = False
            if c.isRental(row['name'], row['invoice_date']):
                isRental = True

            # IF the invoice is marked as cancelled before the therapist is paid for it, then it's not included in their invoice
            if row['associate_invoice_fk'] is None and row['status'] != 'cancelled':
                isNewOrChangedClientInvoice = True
                associate_invoice = None
                if myDictionary.get(collector) is not None:
                    cursor2.execute('''SELECT last_update, id from associate_invoices where id=?''',
                                    (myDictionary[collector],))
                    associate_invoice = cursor2.fetchone()
                if associate_invoice is None:
                    # the client invoice has never been processed.
                    # todayStr = str(today.strftime("%Y%m%d-%H%M%S"))

                    cursor2.execute('''INSERT INTO associate_invoices(name, number, invoice_date, tax, subtotal, total, status, last_update)
                                  VALUES(?,?,?,?,?,?,?,?)''',
                                    (collector, collector + "-" + nowStr, eDateInvoices, 0, 0, 0, 'sent', today))
                    myDictionary[collector] = cursor2.lastrowid
                    invoiceid = myDictionary[collector]

                cursor2.execute('''UPDATE invoices set associate_invoice_fk=? 
                    where number=?''', (myDictionary[collector], row['number']))
                database.commit()
            else:
                cursor2.execute('''SELECT last_update, id from associate_invoices where id=?''',
                                (row['associate_invoice_fk'],))
                # writer = csv.DictWriter(csvwriter, fieldnames=['Customer','Email','Telephone','StartDate','EndDate'])
                associate_invoice = cursor2.fetchone()
                if associate_invoice is not None and not isRental:
                    a = parser.parse(associate_invoice['last_update'])
                    b = parser.parse(row['last_update'])
                    if associate_invoice is not None and a <= b and row['status'] == 'paid':
                        isNewOrChangedClientInvoice = True
                        print 'CHANGE CHANGE in status of ' + row['number']
                        if myDictionary.get(collector) is None:
                            cursor2.execute('''INSERT INTO associate_invoices(name, number, invoice_date, tax, subtotal, total, status, last_update)
                                      VALUES(?,?,?,?,?,?,?,?)''',
                                            (collector, collector + "-" + nowStr, eDateInvoices, 0, 0, 0, 'sent', today))
                            myDictionary[collector] = cursor2.lastrowid
                        invoiceid = myDictionary[collector]
                        cursor2.execute('''UPDATE invoices set associate_invoice_fk=? where id=?''',
                                        (invoiceid, row['id']))
                        # cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update)
                        #               VALUES(?,?,?,?,?)''', (row['number'], associate_invoice['invoice_date'], "AS_INV_FK_CHANGE",
                        #                                     row['associate_invoice_fk'] + ' is changed to '+invoiceid,today))

                        database.commit()
                elif associate_invoice is None and row['prev_associate_invoice_fk'] is not None and row['status'] == 'cancelled':
                    #a = parser.parse(associate_invoice['last_update'])
                    #b = parser.parse(row['last_update'])
                    isNewOrChangedClientInvoice = True
                    print 'CHANE STATUS TO CANCEL for ' + row['number']
                    if myDictionary.get(collector) is None:
                        cursor2.execute('''INSERT INTO associate_invoices(name, number, invoice_date, tax, subtotal, total, status, last_update)
                                  VALUES(?,?,?,?,?,?,?,?)''',
                                        (collector, collector + "-" + nowStr, eDateInvoices, 0, 0, 0, 'sent',
                                         today))
                        myDictionary[collector] = cursor2.lastrowid
                    invoiceid = myDictionary[collector]
                    cursor2.execute('''UPDATE invoices set associate_invoice_fk=? where id=?''',
                                    (invoiceid, row['id']))


                    database.commit()


            # Stop processing already processed invoice
            if isNewOrChangedClientInvoice == False:
                continue

            if isRental:
                client = Client(dict.get(collector))
                provider = Provider(payee, bank_account=' ', bank_code=' ')
            else:
                client = Client(payee)
                provider = Provider(dict.get(collector), bank_account=' ', bank_code=' ')

            cursor4.execute('''SELECT * FROM invoices where id=?''', (row['id'],))
            updatedRow = cursor4.fetchone()

            invoice=createInvoicePDF(client, invoice, invoicePath, invoiceid, isRental, None, isPercentage, provider, updatedRow, sDate, isPartialInvoice)

        if invoiceid is not None and invoice is not None:
            cursor2.execute('''UPDATE associate_invoices set total=?, subtotal=?, tax=? 
                    where id=?''', (float(invoice.price_tax), float(invoice.price), float(invoice.price_tax-invoice.price), invoiceid))
            database.commit()
            print "total price is "+str(invoice.price)
            print "total tax is "+str(invoice.price_tax)

        if invoice is not None:

            pdf = MyInvoice(invoice)

            pdf_name = collector + "-" + provider.address + "-" + str(today.strftime("%Y%m%d-%H%M%S")) + ".pdf";
            pdfPath = invoicePath + "/" + pdf_name
            pdf.gen(pdfPath,
                    generate_qr_code=True)
            file = open(pdfPath, 'r')
            file_content = file.read()
            cursor2.execute('''UPDATE associate_invoices set invoice_file=? 
                    where id=?''', (file_content, invoiceid))
            database.commit()
            gdrive.create_file(gfile, pdf_name, pdfPath)

            file.close()
    except Exception as e:
        print e.message
        # Roll back any change if something goes wrong
        database.rollback()

        raise e
    finally:
        cursor.close()
        cursor2.close()
        cursor3.close()
        cursor4.close()
        cursor5.close()
        # Close the db connection
        database.close()



def createInvoicePDF(client, invoice, invoicePath, invoiceid, isRental, SWorPsy, isPercentage, provider, row, sDate, isPartialInvoice):
    counsellingType = row['service_type']
    noshow = False
    abbrName, invoice, multiplier, noshow = c.getInvoiceRowInfo(client, counsellingType, invoice, invoiceid,
                                                              isPartialInvoice, noshow, provider, row, sDate)

    c.addItem(SWorPsy, abbrName, invoice, isPercentage, isRental, multiplier, noshow, row)
    return invoice


def sendShareLink(list, fromDate, toDate):
    try:
        print 'Sending shared link from: '+fromDate+ " to: "+toDate
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.ehlo()
        to = ['amiralig@gmail.com']

        subject = 'Payroll link from: '+fromDate+', to: '+toDate

        msg = MIMEText(list)
        msg['From'] = sent_from
        msg['To'] = COMMASPACE.join(to)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = subject
        # msg['Content-Type'] = "text/html"

        server.sendmail(sent_from, to, msg.as_string())
        server.close()

    except Exception as e:
        print 'Something went wrong sending link to associates' + e.message


def analyze_old(collector, payee, db):
    try:
        processedFile = sys.argv[1]
        reader = openpyxl.load_workbook(filename=processedFile, data_only=True)
        today = datetime.datetime.now().today().ctime()
        inSheet = reader.get_sheet_by_name(collector)
        # database=sqlite3.connect(db)
        # cursor=database.cursor()
        # cursor.execute('''SELECT * FROM invoices''')

        # writer = csv.DictWriter(csvwriter, fieldnames=['Customer','Email','Telephone','StartDate','EndDate'])
        # writer.writeheader()
        myworkbook = openpyxl.load_workbook(filename=processedFile)
        myDictionary = {}
        superDictionary={}
        invoice = None
        isRental = False
        if collector == 'Dave' or collector == 'Stefanie':
            isRental = True
        for row in inSheet.rows:
            # for row in cursor:
            currentRow = str(row[0].row)
            if currentRow == '1':
                continue

            if inSheet['A' + currentRow].value is None:
                break

            if isRental:
                client = Client(dict.get(collector))
                provider = Provider(payee, bank_account=' ', bank_code=' ')
            else:
                client = Client(payee)
                provider = Provider(dict.get(collector), bank_account=' ', bank_code=' ')

            if invoice is None:
                invoice = Invoice(client, provider, creator)
                invoice.use_tax = True
                invoice.currency_locale = 'en_US.UTF-8'
                invoice.currency = 'CAD'

            if isRental:
                if inSheet['C' + currentRow].value is None or inSheet['C' + currentRow].value == 0:
                    invoice.add_item(Item(1, 0, inSheet['F' + currentRow].value, tax=13))
                elif inSheet['C' + currentRow].value is not None and inSheet['C' + currentRow].value > 130:
                    invoice.add_item(Item(1, 45, inSheet['F' + currentRow].value, tax=13))
                elif inSheet['C' + currentRow].value is not None and inSheet['C' + currentRow].value <= 130:
                    invoice.add_item(Item(1, 30, inSheet['F' + currentRow].value, tax=13))
            else:
                if inSheet['C' + currentRow].value is None or inSheet['C' + currentRow].value == 0:
                    invoice.add_item(Item(1, 0, inSheet['F' + currentRow].value))
                elif inSheet['C' + currentRow].value is not None and inSheet['C' + currentRow].value > 130:
                    invoice.add_item(Item(1, inSheet['C' + currentRow].value - 45, inSheet['F' + currentRow].value,
                                          tax=0 if inSheet['B' + currentRow].value == 0 else 13))
                elif inSheet['C' + currentRow].value is not None and inSheet['C' + currentRow].value <= 130:
                    invoice.add_item(Item(1, inSheet['C' + currentRow].value - 30, inSheet['F' + currentRow].value,
                                          tax=0 if inSheet['B' + currentRow].value == 0 else 13))

            pdf = SimpleInvoice(invoice)
            pdf.gen(invoicePath + "/" + inSheet.title + "-" + today + ".pdf", generate_qr_code=True)

    except Exception as e:
        print e.message
        # Roll back any change if something goes wrong
        database.rollback()
        raise e
    finally:
        # Close the db connection
        database.close()


db = sys.argv[1] + "/lifeVCC"
sDate = sys.argv[2]
eDate = sys.argv[3]
invoicePath = sys.argv[1]
invoicePath = invoicePath+"/"+sDate+"-"+eDate
list_of_files = glob.glob('/Users/amirali/Downloads/*.xlsx')
# = max(list_of_files, key=os.path.getctime)
latest_file = max(list_of_files, key=os.path.getctime)


if not os.path.exists(invoicePath):
    os.makedirs(invoicePath)
else:
     today=datetime.datetime.now().today()
     todayStr=str(today.strftime("%Y%m%d-%H%M%S"))
     invoicePath=invoicePath+todayStr
     os.makedirs(invoicePath)

backup = invoicePath+"/current.xlsx";
backupDb = invoicePath+"/lifeVCC"+eDate
copyfile(latest_file,backup)
copyfile(db,backupDb+"pre")

file = gdrive.create_drive(sDate,eDate)
analyze('Dave', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Stefanie', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Sabrina', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Mariella', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Johanna', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Meri', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Melissa', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Teresa', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Rosemary', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Peter', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Joanne', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
#analyze('Alex', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Alex', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)

analyze('Sandra', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Mary', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Lisa', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
analyze('Helena', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, eDate, invoicePath, file)
#analyze('Kelly-Lee', 'Life in Harmony at Vaughan Counselling Centre', db, sDate, '2019-11-02', invoicePath, file)

copyfile(db,backupDb+"-post")
userList = ['info@lifeinharmony.ca', 'mariellap@lifeinharmony.ca', 'johannag@lifeinharmony.ca', 'sabrinag@lifeinharmony.ca']
gdrive.share_file(userList, file.get('id'))
#sendShareLink(file.get('webViewLink'), sDate, eDate)
