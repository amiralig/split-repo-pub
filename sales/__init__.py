import wget
import openpyxl
import os
import sys
import sqlite3
import glob

from datetime import datetime
from icalendar import Calendar
from shutil import copyfile, copy2

from common import Common

try:
    print "yes"
except ImportError:
    print "yws1"
# url = "http://lifeinharmony.janeapp.com/ical/QdK1sX3uAX1O6g1VkJfS/appointments.ics"
print "hi1"
c = Common()


# testfile = urllib.URLopener()
# testfile.retrieve("http://lifeinharmony.janeapp.com/ical/QdK1sX3uAX1O6g1VkJfS/appointments.ics", "/Users/amirali/Downloads/sabrina-appointments.ics")

def get_calendar(therapist):
    appFile = ""
    if str(therapist).find("Sabrina") != -1:
        appFile = wget.download("http://lifeinharmony.janeapp.com/ical/QdK1sX3uAX1O6g1VkJfS/appointments.ics")
    elif str(therapist).find("Dave") != -1:
        appFile = wget.download("https://lifeinharmony.janeapp.com/ical/VKBUXCMT3aMsfdVL4bOi/appointments.ics")
    # elif str(therapist).find("Stefanie") != -1:
    #    appFile = wget.download()


    g = open(appFile, 'rb')
    gcal = Calendar.from_ical(g.read())

    sDate = sys.argv[1]
    eDate = sys.argv[2]
    fileToWrite = sys.argv[3]

    print "Processing for start date: " + str(sDate)
    print "end date" + str(eDate)
    print "writing to file " + fileToWrite

    stringDate = sDate
    stringDate2 = eDate

    sDate = datetime.strptime(sDate, '%Y-%m-%d')
    myworkbook = openpyxl.load_workbook(filename=fileToWrite)
    worksheet = myworkbook.get_sheet_by_name('Sabrina')
    eDate = datetime.strptime(eDate, '%Y-%m-%d')
    i = 1
    for event in gcal.walk('vevent'):
        date = event['DTSTART'].dt
        date = date.replace(tzinfo=None)
        summery = event.get('summary')
        if date >= sDate and date <= eDate:
            print str(summery)
            print "i is" + str(i)
            s = "B" + str(i)
            s1 = "A" + str(i)
            worksheet[s1].value = date
            worksheet[s].value = summery
            i = i + 1

    myworkbook.save("/Users/amirali/Documents/Life-VCC/lifeVCC-" + stringDate + "-to-" + stringDate2 + ".xlsx")
    g.close()
    os.remove(g.name)


def jane_sales(combinedFile, separatedFile, dbXlsxPath):
    # with open(combinedFile) as csvfile:
    reader = openpyxl.load_workbook(filename=combinedFile)
    db = sqlite3.connect(dbXlsxPath + "/lifeVCC")
    db.text_factory = str

    try:
        cursor = db.cursor()

        print 'table does not exist'
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoices(id INTEGER PRIMARY KEY, name TEXT, client TEXT, number TEXT unique,
                               invoice_date DATETIME, tax REAL, subtotal REAL, total REAL, status TEXT, last_update DATETIME,
                               associate_invoice_fk INTEGER, supervisor_invoice_fk INTEGER, service_type TEXT, is_partial_invoice INTEGER, prev_associate_invoice_fk INTEGER)
        ''')

        cursor.execute('''
                    CREATE TABLE IF NOT EXISTS invoice_audit(id INTEGER PRIMARY KEY,  
                    invoice_number TEXT, invoice_date DATETIME, audit_event TEXT, audit_desc TEXT, last_update DATETIME
                )''')
        cursor.execute('''
                    CREATE TABLE IF NOT EXISTS sales(id INTEGER PRIMARY KEY,  
                    filename TEXT, sales_file BLOB, run_date DATETIME
                )''')

        inSheet = reader.get_sheet_by_name('Export')
        # writer = csv.DictWriter(csvwriter, fieldnames=['Customer','Email','Telephone','StartDate','EndDate'])
        # writer.writeheader()
        myworkbook = openpyxl.load_workbook(filename=separatedFile)
        today = datetime.now()

        file = open(combinedFile, 'r')
        file_content = file.read()
        file.close()
        cursor.execute('''INSERT INTO sales (filename, sales_file, run_date) values(?,?,?)''',
                       (combinedFile, file_content, today))
        db.commit()

        myDictionary = {}

        newInvoices = []
        for row in inSheet.rows:
            currentRow = str(row[0].row)
            # outputRow = str(currentRow + 1)
            if currentRow == '1':
                continue
            # print(row['Email'])
            if inSheet['A' + currentRow].value is None:
                break
            firstName = inSheet['G' + currentRow].value
            firstName = str(firstName).split()[0]

            if firstName in myDictionary:
                value = myDictionary[firstName]
                value += 1
                myDictionary[firstName] = value
                # print myDictionary[eachItem]
            else:
                myDictionary[firstName] = 2

            counsellingType = inSheet['F' + currentRow].value
            subtotal = inSheet['M' + currentRow].value
            hst = inSheet['N' + currentRow].value
            total = inSheet['O' + currentRow].value
            invoiceno = inSheet['I' + currentRow].value
            patient = inSheet['D' + currentRow].value
            status = inSheet['L' + currentRow].value
            # worksheet = myworkbook.get_sheet_by_name(firstName)
            # if worksheet
            date = inSheet['B' + currentRow].value

            i = LifeInvoice(firstName, patient, date, invoiceno, status, subtotal, hst, total, counsellingType)
            newInvoices.append(i)


        for eachInvoice in newInvoices:


            todayStr = str(today.strftime("%Y%m%d-%H%M%S"))
            counsellingType = eachInvoice.counsellingType
            subtotal = eachInvoice.subtotal
            hst = eachInvoice.hst
            total = eachInvoice.total
            invoiceno = eachInvoice.invoiceno
            patient = eachInvoice.patient
            status = eachInvoice.status

            date = eachInvoice.date
            firstName = eachInvoice.firstname
            # worksheet['A' + str(myDictionary[firstName])].value = date
            # worksheet['C' + str(myDictionary[firstName])].value = subtotal
            # worksheet['B' + str(myDictionary[firstName])].value = hst
            # worksheet['D' + str(myDictionary[firstName])].value = total
            # worksheet['E' + str(myDictionary[firstName])].value = invoiceno
            # worksheet['F' + str(myDictionary[firstName])].value = patient
            # worksheet['']

            cursor2 = db.cursor()
            cursor2.execute(
                '''SELECT status, number, id, service_type, tax, subtotal, total, client, associate_invoice_fk, invoice_date from invoices where number=?''',
                [invoiceno])
            print 'analyzing' + invoiceno

            oneinvoice = cursor2.fetchone()
            if oneinvoice is not None and (
                            str(oneinvoice[0]) != status or counsellingType !=
                        oneinvoice[3]):
                print "status is " + status
                print "old status is " + oneinvoice[0]
                print 'CHANGE CHANGE in the content of :' + invoiceno

                if oneinvoice[5] != subtotal and status == 'paid' and oneinvoice[0] == 'paid':
                    cursor2.execute(
                        '''UPDATE invoices set status=?, last_update=?, where id=?''',
                        ('cancelled', today, oneinvoice[2]))

                    cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                    VALUES(?,?,?,?,?)''', (
                        invoiceno, date, "INV_NO_CHANGE", "old no is " + oneinvoice[1] + " - new no is " + invoiceno,
                        today))
                elif str(oneinvoice[0]) != status:
                    cursor2.execute(
                        '''UPDATE invoices set status=?, last_update=?, tax=?, subtotal=?, total=? where id=?''',
                        (status, todayStr, hst, subtotal, total, oneinvoice[2]))
                    cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                    VALUES(?,?,?,?,?)''', (invoiceno, date, "INV_STATUS_CHANGE",
                                                           "old status is " + oneinvoice[
                                                                                                   0] + " - new status is " + status,
                                                           today))
                if oneinvoice[3] != counsellingType and not (
                        str(counsellingType).lower() == 'counselling - fee only' and status == 'paid'):
                    cursor2.execute('''UPDATE invoices set service_type=?, last_update=? where id=?''',
                                    (counsellingType, todayStr, oneinvoice[2]))
                    cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                    VALUES(?,?,?,?,?)''', (invoiceno, date, "INV_SERV_TYPE_CHANGE",
                                                           "old type is " + oneinvoice[
                                                                                                   3] + " - new type is " + counsellingType,
                                                           today))


            elif oneinvoice is not None and str(oneinvoice[0]) == status:
                print 'No change in status of invoice:' + invoiceno
            else:
                print 'analyzibbbng' + invoiceno
                # This could be a change of invoice no, we have to find out why invoice no changed.
                # Let's get the invoice with the same date and time
                cursor2.execute(
                    '''SELECT status, number, id, service_type, tax, 
                        subtotal, total, client, associate_invoice_fk, invoice_date from invoices where invoice_date=? and name=? order by id desc''',
                    (date, firstName))
                oneinvoice2 = cursor2.fetchone()

                if oneinvoice2 is not None and oneinvoice2[1] != invoiceno:
                    # Invoice number is changed but the invoice is already paid out  and the price matches
                    if oneinvoice2[7] == patient and oneinvoice2[5] == subtotal and oneinvoice2[8] is not None and \
                                    oneinvoice2[0] == 'paid':
                        cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                                            VALUES(?,?,?,?,?)''', (
                            invoiceno, date, "INV_NO_NAME_CHANGE",
                            "ignore - no name different but old no is " + oneinvoice2[1] + " - new no is " + invoiceno,
                            today))
                        print invoiceno + ':1'


                    elif oneinvoice2[0] == 'paid' and oneinvoice2[9] > '2019-04-01':
                            #cursor2.execute(
                            #    '''UPDATE invoices set status=?, last_update=? where id=?''',
                            #    ('cancelled', today, oneinvoice2[2]))

                            oldInvoiceFound = False
                            if oneinvoice2[7] != patient:
                                for j in newInvoices:
                                    if oneinvoice2[7] == j.patient and oneinvoice2[9] == j.date:
                                        # the patient still exists for that date, it is a double booking
                                        oldInvoiceFound = True
                            if oldInvoiceFound is False:
                                # it is not a double booking but a client name change, cancel the old invoice and add the new one
                                if oneinvoice2[8] is not None:
                                    # Insert a cancel record if the invoice is already paid to the therapist
                                    cursor2.execute('''INSERT INTO invoices(name, client, number, invoice_date, tax, subtotal, total, status, last_update, service_type, prev_associate_invoice_fk)
                                                  VALUES(?,?,?,?,?,?,?,?,?,?,?)''', (
                                        firstName, patient, oneinvoice2[1]+'cancel', date, oneinvoice2[4], oneinvoice2[5], oneinvoice2[6], 'cancelled', today,
                                        oneinvoice2[3], oneinvoice2[8]))
                                else:
                                    #otherwise cancel existing one
                                    cursor2.execute('''update invoices set status=?, last_update=? where id=?''',('cancelled', today, oneinvoice2[2]))

                                cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                                                    VALUES(?,?,?,?,?)''', (
                                    invoiceno, date, "INV_PREV_CANCEL_NEW_CREATE_GEN2",
                                    "cancelled no/name " + oneinvoice2[1] + "/" + oneinvoice2[
                                        7] + " - new no/name " + invoiceno + "/" + patient,
                                    today))
                                cursor2.execute('''INSERT INTO invoices(name, client, number, invoice_date, tax, subtotal, total, status, last_update, service_type)
                                              VALUES(?,?,?,?,?,?,?,?,?,?)''', (
                                    firstName, patient, invoiceno, date, hst, subtotal, total, status, today,
                                    counsellingType))
                            else:
                                cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                                                        VALUES(?,?,?,?,?)''', (
                                    invoiceno, date, "INV_DOUBLE_BOOKING_GEN2",
                                    "name is different, old no/name is " + oneinvoice2[1] + "/" + oneinvoice2[
                                        7] + " - new no/name is " + invoiceno + "/" + patient,
                                    today))
                                cursor2.execute('''INSERT INTO invoices(name, client, number, invoice_date, tax, subtotal, total, status, last_update, service_type)
                                              VALUES(?,?,?,?,?,?,?,?,?,?)''', (
                                    firstName, patient, invoiceno, date, hst, subtotal, total, status, today,
                                    counsellingType))
                                print invoiceno + 'double booking:2'

                    elif oneinvoice2[8] is not None and c.isRental(firstName, oneinvoice2[9]):
                        cursor2.execute('''INSERT INTO invoice_audit(invoice_number, invoice_date, audit_event, audit_desc, last_update) 
                                                            VALUES(?,?,?,?,?)''', (
                            invoiceno, date, "INV_NO_NAME_CHANGE",
                            "ignore rental - old no is " + oneinvoice2[1] + " - new no is " + invoiceno,
                            today))
                        print invoiceno + ':3'

                    else:
                        # The previous one is cancelled, so we have to take into account this new one
                        cursor2.execute('''INSERT INTO invoices(name, client, number, invoice_date, tax, subtotal, total, status, last_update, service_type, is_partial_invoice)
                                          VALUES(?,?,?,?,?,?,?,?,?,?,?)''', (
                            firstName, patient, invoiceno, date, hst, subtotal, total, status, today,
                            counsellingType, 0))
                        print invoiceno + ':4'

                else:
                    # newStr = datetime.strptime(date, '%Y-%m-%d %H:%M:%S%f%z').strftime('%Y%m%d-%H%M%S')
                    cursor2.execute('''INSERT INTO invoices(name, client, number, invoice_date, tax, subtotal, total, status, last_update, service_type)
                                  VALUES(?,?,?,?,?,?,?,?,?,?)''', (
                        firstName, patient, invoiceno, date, hst, subtotal, total, status, today, counsellingType))
                    print invoiceno + ':4'

            db.commit()
            # myworkbook.save(dbXlsxPath + "-" + todayStr + ".xlsx")

    except Exception as e:
        print e.message
        # Roll back any change if something goes wrong
        db.rollback()
        raise e
    finally:
        # Close the db connection
        cursor.close()
        cursor2.close()
        db.close()

class LifeInvoice:
  def __init__(self, firstname, patient, date, invoiceno, status, subtotal, hst, total, counsellingType):
    self.firstname = firstname
    self.counsellingType = counsellingType
    self.subtotal = subtotal
    self.hst = hst
    self.total = total
    self.invoiceno = invoiceno
    self.patient = patient
    self.status = status
    # worksheet = myworkbook.get_sheet_by_name(firstName)
    # if worksheet
    self.date = date


list_of_files = glob.glob('/Users/amirali/Public/reports/*.xlsx')

latest_file = max(list_of_files, key=os.path.getctime)
copy2(latest_file, '/Users/amirali/Downloads/')

print latest_file
# latest_file = "/Users/Amirali/Documents/Life-VCC/payroll/2019-02-08-2019-02-22/Sales_20190201_20190223(13).xlsx"

jane_sales(latest_file, sys.argv[1], sys.argv[2])
# jane_sales(path, sys.argv[1], sys.argv[2])
